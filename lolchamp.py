import PySimpleGUI as sg
import os.path
import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl import Workbook
from table import create_table

champbp = []
champwp = []
wb= Workbook()
ws=wb.active

def champfinder(x, y):
    x = champbp1
    y = champwp1
    data_file = 'matchups.xlsx'
    wb = load_workbook(data_file)

    # Load one worksheet.
    ws = wb[champbp1]
    all_rows = list(ws.rows)

    for cell in all_rows[0]:
        row1 = cell.value

sg.theme('DarkAmber')

layout = [
    [sg.Text('Enemy Champ', size =(15, 1)), sg.InputText(do_not_clear=False, key='echamp')],
    [sg.Submit(), sg.Cancel()]
]

window = sg.Window('Window Title', layout)

while True:
    event, values = window.read()
    if event == sg.WIN_CLOSED or event == 'Cancel': # if user closes window or clicks cancel
        break
    if event == "Submit":
        champ = values['echamp']
        champbp = pd.read_excel('matchups.xlsx', sheet_name=champ)
        champwp = pd.read_excel('matchups.xlsx', sheet_name=champ)
        sg.popup('Matchups', champbp)
window.close()
